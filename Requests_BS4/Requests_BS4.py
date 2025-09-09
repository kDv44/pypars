import re
import requests

from bs4 import BeautifulSoup
from openpyxl import Workbook


headers = {"User-Agent": "Mozilla/5.0"}
url = "https://brain.com.ua/ukr/Mobilniy_telefon_Apple_iPhone_16_Pro_Max_256GB_Black_Titanium-p1145443.html"
#url = "https://brain.com.ua/ukr/Mobilniy_telefon_Xiaomi_Redmi_Note_14_8_256GB_Midnight_Black_1123261-p1190652.html" #for tests
#url = "https://brain.com.ua/ukr/Mobilniy_telefon_Xiaomi_Redmi_Note_14_8_256GB_Mist_Purple_1123263-p1190654.html" #for tests


response = requests.get(url)

soup  = BeautifulSoup(response.text, "lxml")

product = {}

wb = Workbook()
ws = wb.active
ws.title = "Product"

try:
    product["Full name"] = soup.find("span", string="Модель").find_next_sibling("span").text.strip()
except AttributeError:
    product["Full name"] = None

try:
    product["Color"] = soup.find("a", title=lambda t: t and "Колір" in t).text.strip()
except AttributeError:
    product["Color"] = None

try:
    product["Memory size"] = soup.find("a", title=lambda t: t and "Вбудована пам'ять" in t).text.strip()
except AttributeError:
    product["Memory sizes"] = None

# not finder this
# try:
#     product["Salesperson"] = soup.find('').text.strip()
# except AttributeError:
#     product["Salesperson"] = None

try:
    product["Price"] =  " ".join(soup.find("div", class_="br-pr-price main-price-block").text.split())
except AttributeError:
    product["Price"] = None

try:
    product["Red price"] = soup.find('span', class_="red-price").text.strip()
except AttributeError:
    product["Red price"] = None

try:
    img_links = soup.select("div.product-block-right img.br-main-img")
    product["Photo"] = [img["src"] for img in img_links if "src" in img.attrs]
except AttributeError:
    product["Photo"] = None

try:
    product["id"] = soup.find("span", class_="br-pr-code-val").text
except AttributeError:
    product["id"] = None

try:
    reviews = "".join(soup.find("a", class_="scroll-to-element brackets-reviews").text.strip())
    product["Number of reviews"] = re.search(r'\d+', reviews).group()
except AttributeError:
    product["Number of reviews"] = None

try:
    product["Screen diagonal"] = soup.find("a", title=lambda t: t and "Діагональ екрану" in t).text.strip()
except AttributeError:
    product["Screen diagonal"] = None

try:
    product["Display resolution"] = soup.find("a", title=lambda t: t and "Роздільна здатність екрану" in t).text.strip()
except AttributeError:
    product["Display resolution"] = None

def clean(text):
    text = " ".join(text.stripped_strings)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s*,\s*", ", ", text)
    return text.strip()

try:
    specs = {}
    for block in soup.select(".br-pr-chr-item"):
        section = block.h3.get_text(strip=True)
        rows = {}
        for row in block.select("div > div"):
            spans = row.find_all("span")
            if len(spans) >= 2:
                key, value = clean(spans[0]), clean(spans[1])
                if "," in value:
                    value = [v.strip() for v in value.split(",")]
                rows[key] = value
        if rows:
            specs[section] = rows

    product["Characteristics"] = {"Characteristics": specs}
except AttributeError:
    product["Characteristics"] = None


print(product)

ws["B2"] = product.get("Full name")

row = 4
specs = product["Characteristics"]["Characteristics"]


for section, values in specs.items():
    ws.cell(row=row, column=1, value=section)
    row += 1
    for key, val in values.items():
        if isinstance(val, list):
            val = ", ".join(val)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=3, value=val)
        row += 1
    row += 1

row += 2
ws.cell(row=row, column=1, value="Photo:")
row += 1

for img_url in product.get("Photo", []):
    ws.cell(row=row, column=2, value=img_url)
    row += 1

wb.save("req_bs4.xlsx")
