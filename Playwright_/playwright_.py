import time

from openpyxl import Workbook, load_workbook


from playwright.sync_api import sync_playwright


product = {}

with sync_playwright() as p:
    browser = p.webkit.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()

    page.goto("https://brain.com.ua/ukr/", timeout=30000)

    search_box = page.locator(".quick-search-input:visible")
    search_box.fill("Apple iPhone 15 128GB Black")
    search_box.press("Enter")

    page.wait_for_selector("//div[contains(@class,'product-wrapper')]", timeout=30000)

    first_link = page.locator("(//div[contains(@class,'product-wrapper')]//a)[1]")
    first_link.wait_for(state="visible", timeout=30000)

    phone_link = first_link.get_attribute("href")
    page.goto(phone_link, timeout=30000)

    try:
        product["Full name"] = page.locator("(//div[contains(@class, 'br-pr-chr-item')][11]//span)[4]").inner_text()
    except AttributeError:
        product["Full name"] = None

    try:
        color = page.locator("(//div[contains(@class, 'br-pr-chr-item')][10]//span)[6]")
        product["color"] = color.inner_text().strip()
    except AttributeError:
        product["color"] = None

    try:
        memory = page.locator("//div[contains(@class, 'br-pr-chr-item')]//a[contains(@title, \"Вбудована пам'ять\")]")
        product["Memory size"] = memory.inner_text().strip()
    except AttributeError:
        product["Memory size"] = None

    try:
        img_elems = page.locator("div.product-block-right img.br-main-img")
        count = img_elems.count()
        product["Photo"] = [img_elems.nth(i).get_attribute("src") for i in range(count) if
                            img_elems.nth(i).get_attribute("src")]
    except AttributeError:
        product["Photo"] = None

    try:
        product_id = page.locator("//span[contains(@class, 'br-pr-code-val')]").first
        product["id"] = product_id.inner_text().strip()
    except AttributeError:
        product["id"] = None

    try:
        reviews =  page.locator("//a[contains(@class, 'forbid-click')]")
        product["Number of reviews"] = reviews.inner_text().strip()
    except AttributeError:
        product["Number of reviews"] = None

    try:
        screen_diagonal = page.locator("(//div[contains(@class, 'br-pr-chr-item')][2]//span)[4]")
        product["Screen diagonal"] = screen_diagonal.inner_text().strip()
    except AttributeError:
        product["Screen diagonal"] = None

    try:
        display_resolution = page.locator("(//div[contains(@class, 'br-pr-chr-item')][2]//span)[6]")
        product["Display resolution"] = display_resolution.inner_text().strip()
    except AttributeError:
        product["Display resolution"] = None


    specs_dict = {}
    spec_blocks = page.locator("xpath=//div[contains(@class,'br-pr-chr-item')]")
    for i in range(spec_blocks.count()):
        block = spec_blocks.nth(i)
        rows = block.locator("xpath=.//div[span]")
        for j in range(rows.count()):
            row = rows.nth(j)
            key_span = row.locator("xpath=.//span[1]")
            value_span = row.locator("xpath=.//span[2]")

            key = key_span.inner_text().strip() if key_span.count() > 0 else None
            value = value_span.inner_text().strip() if value_span.count() > 0 else None

            if key and value:
                specs_dict[key] = value


    product["Specs"] = specs_dict

    print(product)


def write_product_to_excel(product, filename="playwright_output.xlsx"):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    row = ws.max_row + 1 if ws.max_row > 1 else 1

    def write_dict(d, start_row):
        current_row = start_row
        for key, value in d.items():
            if isinstance(value, dict):
                ws.cell(row=current_row, column=1, value=f"{key}:")
                current_row += 1
                current_row = write_dict(value, current_row)
            elif isinstance(value, list):
                ws.cell(row=current_row, column=1, value=f"{key}:")
                current_row += 1
                for item in value:
                    ws.cell(row=current_row, column=2, value=item)
                    current_row += 1
            else:
                ws.cell(row=current_row, column=1, value=key)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
        return current_row

    write_dict(product, row)

    wb.save(filename)


write_product_to_excel(product)
