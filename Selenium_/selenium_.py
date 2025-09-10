import re
import time

from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service



service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)

driver.get("https://brain.com.ua/ukr/")


product = {}


input_search = driver.find_element(By.CLASS_NAME, "quick-search-input")

time.sleep(0)

input_search.send_keys("Apple iPhone 15 128GB Black")
input_search.send_keys(Keys.ENTER)

first_product = driver.find_element(By.XPATH, "(//div[contains(@class, 'product-wrapper')]//a)[1]")
first_product.click()

time.sleep(2)

try:
    name = driver.find_element(By.XPATH, "(//div[contains(@class, 'br-pr-chr-item')][11]//span)[4]")
    product["Full name"] = name.get_attribute("textContent").strip()
except AttributeError:
    product["Full name"] = None

try:
    color = driver.find_element(By.XPATH, "(//div[contains(@class, 'br-pr-chr-item')][10]//span)[6]")
    product["Color"] = color.get_attribute("textContent").strip()
except AttributeError:
    product["Color"] = None

try:
    memory = driver.find_element(By.XPATH, "//div[contains(@class, 'br-pr-chr-item')]//a[contains(@title, \"Вбудована пам'ять\")]")
    product["Memory size"] = memory.get_attribute("textContent").strip()
except AttributeError:
    product["Memory size"] = None

try:
    product["Price"] = " ".join(driver.find_element(By.XPATH, "(//div[contains(@class, 'br-pr-np')])[1]").text.split())
except AttributeError:
    product["Price"] = None

try:
    img_elems = driver.find_elements(By.CSS_SELECTOR, "div.product-block-right img.br-main-img")
    product["Photo"] = [img.get_attribute("src") for img in img_elems if img.get_attribute("src")]
except AttributeError:
    product["Photo"] = None

try:
    product_id = driver.find_element(By.XPATH, "//span[contains(@class, 'br-pr-code-val')]")
    product["id"] = product_id.get_attribute("textContent").strip()
except AttributeError:
    product["id"] = None

try:
    reviews = driver.find_element(By.XPATH, "//a[contains(@class, 'forbid-click')]")
    product["Number of reviews"] = reviews.get_attribute("textContent").strip()
except AttributeError:
    product["Number of reviews"] = None

try:
    screen_diagonal = driver.find_element(By.XPATH, "(//div[contains(@class, 'br-pr-chr-item')][2]//span)[4]")
    product["Screen diagonal"] = screen_diagonal.get_attribute("textContent").strip()
except AttributeError:
    product["Screen diagonal"] = None

try:
    display_resolution = driver.find_element(By.XPATH, "(//div[contains(@class, 'br-pr-chr-item')][2]//span)[6]")
    product["Display resolution"] = display_resolution.get_attribute("textContent").strip()
except AttributeError:
    product["Display resolution"] = None

def clean(text):
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s*,\s*", ", ", text)
    return text.strip()

specs = {}

spec_blocks = driver.find_elements(By.XPATH, "//div[contains(@class,'br-pr-chr-item')]")

for block in spec_blocks:
    rows = block.find_elements(By.XPATH, ".//div[span]")

    for row in rows:
        spans = row.find_elements(By.TAG_NAME, "span")

        if len(spans) >= 2:
            key = spans[0].text.strip()
            value = spans[1].text.strip()
            links = spans[1].find_elements(By.TAG_NAME, "a")

            if links:
                link_texts = [a.text.strip() for a in links if a.text.strip()]
                value = " , ".join(link_texts)

            specs[key] = value


product["Characteristics"] = specs

driver.quit()

print(product)


def write_product_to_excel(product, filename="selenium_output.xlsx"):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    row = ws.max_row + 1 if ws.max_row > 1 else 1

    def write_dict(d, start_row):
        current_row = start_row
        for key, value in d.items():
            if isinstance(value, dict):
                ws.cell(row=current_row, column=1, value=f"{key}:")
                current_row += 1
                current_row = write_dict(value, current_row)
            elif isinstance(value, list):
                ws.cell(row=current_row, column=1, value=f"{key}:")
                current_row += 1
                for item in value:
                    ws.cell(row=current_row, column=2, value=item)
                    current_row += 1
            else:
                ws.cell(row=current_row, column=1, value=key)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
        return current_row

    write_dict(product, row)

    wb.save(filename)


write_product_to_excel(product)

